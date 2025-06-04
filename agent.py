# Standard library imports
import os
import json
import logging
import base64
import openpyxl
from datetime import datetime
from dataclasses import dataclass
from io import BytesIO
from typing import Dict, List, Optional, Any, Union

# Third-party imports
import PyPDF2
from dotenv import load_dotenv
from google.adk.agents import LlmAgent, SequentialAgent
from google.oauth2.service_account import Credentials as ServiceAccountCredentials
from googleapiclient.discovery import build, Resource
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Configuration constants
MODEL_NAME: str = os.getenv("MODEL", "gemini-2.0-flash")
CREDENTIALS_PATH: str = os.getenv(
    "GOOGLE_CREDENTIALS_PATH", "./gen-lang-client-0974464525-ba886ad7f95a.json"
)


@dataclass
class GuidelineInfo:
    """
    Information about a guideline document.

    Attributes:
        id (str): Unique identifier for the guideline
        name (str): Name of the guideline document
        content (str): Content of the guideline document
        last_updated (str): Last update timestamp
        type (str): Type of guideline (e.g., "review", "amendment", "general")
    """

    id: str
    name: str
    content: str
    last_updated: str
    type: str


class GoogleDriveService:
    """
    Service class for Google Drive operations.
    Handles interactions with Google Drive API for document management.
    """

    def __init__(self, credentials_json: Optional[str] = None) -> None:
        """
        Initialize Google Drive service with credentials.
        Only sets up the service connection.
        """
        self.service = self._initialize_service(credentials_json or CREDENTIALS_PATH)
        self.guidelines: Dict[str, GuidelineInfo] = {}

    def _initialize_service(self, credentials_json: str) -> Resource:
        """
        Initialize Google Drive service with credentials.

        Args:
            credentials_json (str): Path to credentials JSON file or JSON string

        Returns:
            googleapiclient.discovery.Resource: Initialized Google Drive service

        Raises:
            ValueError: If no credentials provided
            FileNotFoundError: If credentials file not found
            json.JSONDecodeError: If credentials file is invalid JSON
            Exception: If service initialization fails
        """
        try:
            if not credentials_json:
                raise ValueError("No credentials provided")

            # Load credentials from file
            with open(credentials_json, "r") as f:
                credentials_info = json.load(f)

            credentials = ServiceAccountCredentials.from_service_account_info(
                credentials_info,
                scopes=["https://www.googleapis.com/auth/drive.readonly"],
            )

            return build("drive", "v3", credentials=credentials)

        except FileNotFoundError:
            logger.error(f"Credentials file not found at {credentials_json}")
            raise
        except json.JSONDecodeError:
            logger.error(f"Invalid JSON in credentials file at {credentials_json}")
            raise
        except Exception as e:
            logger.error(f"Failed to initialize Google Drive service: {e}")
            raise

    def load_guidelines(self, tool_context) -> str:
        """
        Load guideline documents from Google Drive.
        Uses ADK session state for caching.

        Args:
            tool_context: ADK ToolContext for session state access

        Returns:
            str: Status message about loaded guidelines
        """
        try:
            # Check session state for cache
            state = tool_context.state # Access state via tool_context
            cache_key = 'guidelines_cache'
            cache_time = state.get('guidelines_cache_time')
            
            # Use cache if available and not expired (1 hour)
            if cache_key in state and cache_time:
                if datetime.now().timestamp() - cache_time < 3600:  # 1 hour
                    self.guidelines = state[cache_key]
                    # Add guideline objects back to the instance dictionary
                    for key, value in self.guidelines.items():
                         if isinstance(value, dict):
                              # Recreate GuidelineInfo object from dictionary if needed (depending on how it was stored)
                              self.guidelines[key] = GuidelineInfo(**value)
                    
                    return f"Using cached guidelines (loaded {int((datetime.now().timestamp() - cache_time)/60)} minutes ago)"

            # Search for guideline documents
            search_query = (
                "name contains 'Guideline' or "
                "name contains 'NDA' or "
                "name contains 'PROCUREMENT' or "
                "name contains 'Rules' or "
                "name contains 'Contract' or "
                "name contains 'Act' or "
                "name contains 'Reviewing' or "
                "name contains 'Amendment' or "
                "name contains 'Original'"
            )

            results = (
                self.service.files()
                .list(
                    q=search_query,
                    pageSize=50,
                    fields="files(id,name,mimeType,modifiedTime)",
                )
                .execute()
            )

            files = results.get("files", [])
            if not files:
                return "No guideline documents found in Google Drive."

            # Load guidelines
            self.guidelines.clear()
            loaded_count = 0
            
            guidelines_to_cache = {}

            for file in files:
                try:
                    content = self.get_file_content(file["id"])
                    
                    # --- Modified: Check for error in content before creating GuidelineInfo ---
                    if content.startswith("[Error reading file:") or content.startswith("[Error exporting Google Apps file:") or content.startswith("[Error during OCR:") or content.startswith("[Error: Required libraries for OCR not installed.]"): # Check for known error prefixes
                        logger.warning(f"Skipping guideline file {file['name']} due to content error: {content}")
                        continue # Skip this file if content has an error
                    # --- End Modified ---

                    guideline = GuidelineInfo(
                        id=file["id"],
                        name=file["name"],
                        content=content,
                        last_updated=file.get("modifiedTime", ""),
                        type=self._determine_guideline_type(file["name"]),
                    )
                    self.guidelines[file["id"]] = guideline
                    guidelines_to_cache[file["id"]] = guideline.__dict__ # Prepare for caching
                    loaded_count += 1
                except Exception as e:
                    logger.error(f"Error processing guideline file {file['name']}: {e}")
                    continue

            # Update cache in session state (Store as dictionary for pickling/serialization safety)
            # --- Modified: Cache only successfully loaded guidelines ---
            state[cache_key] = guidelines_to_cache
            # --- End Modified ---
            state['guidelines_cache_time'] = datetime.now().timestamp()

            return f"Successfully loaded {loaded_count} guideline documents."

        except Exception as e:
            logger.error(f"Error loading guidelines: {e}")
            return f"Error loading guidelines: {str(e)}"

    def _determine_guideline_type(self, filename: str) -> str:
        """
        Determine the type of guideline based on filename.

        Args:
            filename (str): Name of the file

        Returns:
            str: Type of guideline ('review', 'amendment', or 'general')
        """
        filename_lower = filename.lower()
        if "review" in filename_lower:
            return "review"
        elif "amendment" in filename_lower:
            return "amendment"
        return "general"

    def get_guideline_content(self, tool_context, guideline_type: Optional[str] = None) -> str:
        """
        Get content of guidelines, optionally filtered by type.
        Will load guidelines if not in cache.

        Args:
            tool_context: ADK ToolContext for session state access
            guideline_type (Optional[str]): Type of guidelines to retrieve

        Returns:
            str: Formatted guideline content
        """
        # Load guidelines if not in cache (pass tool_context)
        if not self.guidelines:
            status = self.load_guidelines(tool_context)
            if "Error" in status:
                return f"Error loading guidelines: {status}"

        response = "Available Guidelines:\n\n"
        
        # --- Modified: Only include guidelines with valid content ---
        valid_guidelines = [g for g in self.guidelines.values() if not (
            g.content.startswith("[Error reading file:") or 
            g.content.startswith("[Error exporting Google Apps file:") or 
            g.content.startswith("[Error during OCR:") or 
            g.content.startswith("[Error: Required libraries for OCR not installed.]")
        )]

        if not valid_guidelines:
             return "No valid guidelines loaded or available. Please try loading guidelines again."

        for guideline in valid_guidelines:
        # --- End Modified ---
            if guideline_type and guideline.type != guideline_type: # Check type here
                continue

            response += (
                f"{guideline.name}\n"
                f"   Last Updated: {guideline.last_updated[:10]}\n"
                f"   Type: {guideline.type}\n"
                f"   Content Preview: {guideline.content[:200]}...\n\n"
            )

        return response

    def list_nda_files(self, folder_id: Optional[str] = None) -> str:
        """
        List all NDA-related files from Google Drive.

        Returns:
            str: Formatted string containing list of NDA files with details
        """
        try:
            # Search for NDA documents with various keywords
            search_query = (
                "name contains 'NDA' or "
                "name contains 'Non-Disclosure' or "
                "name contains 'confidential'"
            )

            # Add folder specific search if folder_id is provided
            if folder_id:
                search_query = f"{search_query} and '{folder_id}' in parents"

            results = (
                self.service.files()
                .list(
                    q=search_query,
                    pageSize=50,
                    fields="files(id,name,mimeType,modifiedTime,size,owners,webViewLink)",
                )
                .execute()
            )

            files = results.get("files", [])

            if not files:
                return "No NDA documents found in Google Drive."

            response = "Found NDA documents in Google Drive:\n\n"
            for i, file in enumerate(files, 1):
                response += (
                    f"{i}. {file['name']}\n"
                    f"   Modified: {file.get('modifiedTime', 'Unknown')[:10]}\n"
                    f"   Link: {file.get('webViewLink', 'No link available')}\n"
                    f"   Type: {file.get('mimeType', 'Unknown')}\n"
                    f"   Size: {self._format_size(int(file.get('size', 0)))}\n\n"
                )

            response += (
                "\nTo view a file's content, use the command: 'view file [number]'"
            )
            return response

        except Exception as e:
            logger.error(f"Error listing NDA files: {e}")
            return f"Error accessing Google Drive: {str(e)}"

    def _format_size(self, size_bytes: int) -> str:
        """
        Format file size in bytes to human readable format.

        Args:
            size_bytes (int): Size in bytes

        Returns:
            str: Formatted size string (e.g., "1.5 MB")
        """
        if size_bytes == 0:
            return "0 B"

        size_names = ("B", "KB", "MB", "GB", "TB")
        i = 0
        while size_bytes >= 1024 and i < len(size_names) - 1:
            size_bytes /= 1024
            i += 1

        return f"{size_bytes:.2f} {size_names[i]}"

    def get_file_content(self, file_id: str) -> str:
        """
        Get content of a file from Google Drive.

        Args:
            file_id (str): ID of the file to retrieve

        Returns:
            str: File content or error message

        Raises:
            Exception: If file content cannot be retrieved
        """
        try:
            # Get file metadata
            file_metadata = self.service.files().get(fileId=file_id).execute()
            mime_type = file_metadata.get("mimeType", "")

            # Differentiate between native Google Workspace files and others
            if mime_type.startswith("application/vnd.google-apps."):
                # Native Google Docs, Sheets, Slides, etc. Use export.
                logger.info(f"Exporting Google Workspace file {file_id} with mimeType {mime_type}")
                return self._export_google_apps_content(file_id, mime_type)
            elif mime_type == "application/pdf":
                 logger.info(f"Getting PDF content for {file_id}")
                 return self._get_pdf_content(file_id) # Already uses get_media
            elif mime_type in [
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
                "application/vnd.ms-excel",  # .xls
            ]:
                 logger.info(f"Getting Excel content for {file_id} with mimeType {mime_type} via openpyxl")
                 return self._get_excel_content_via_openpyxl(file_id) # Already uses get_media
            else:
                # Other file types (uploaded docs, plain text, etc.). Use get_media.
                 logger.info(f"Getting generic file content for {file_id} with mimeType {mime_type} via get_media")
                 return self._get_other_content(file_id) # Uses get_media

        except Exception as e:
            logger.error(f"Error getting file content: {e}")
            return f"[Error reading file: {str(e)}]"

    def _export_google_apps_content(self, file_id: str, mime_type: str) -> str:
        """
        Export content from a Google Apps file (Docs, Sheets, etc.).

        Args:
            file_id (str): ID of the Google Apps file
            mime_type (str): MIME type of the file

        Returns:
            str: Exported content
        """
        try:
            if "document" in mime_type:
                export_mime_type = "text/plain"
            elif "spreadsheet" in mime_type:
                export_mime_type = "text/csv"
            # Add more mime types if needed (e.g., "text/html" for Slides)
            else:
                export_mime_type = "text/plain" # Default export type

            content = (
                self.service.files()
                .export(fileId=file_id, mimeType=export_mime_type)
                .execute()
            )
            return content.decode("utf-8", errors="ignore")

        except HttpError as error:
            logger.error(f"HTTP Error exporting Google Apps file {file_id}: {error}")
            return f"[Error exporting Google Apps file: {error}]"
        except Exception as e:
            logger.error(f"Error exporting Google Apps file {file_id}: {e}")
            return f"[Error exporting Google Apps file: {e}]"

    def _get_pdf_content(self, file_id: str) -> str:
        """
        Get content from a PDF file.

        Args:
            file_id (str): ID of the PDF file

        Returns:
            str: Extracted text from PDF
        """
        request = self.service.files().get_media(fileId=file_id)
        fh = BytesIO()
        downloader = MediaIoBaseDownload(fh, request)

        done = False
        while not done:
            _, done = downloader.next_chunk()

        fh.seek(0)
        text = ""
        try:
            # Try extracting text using PyPDF2
            reader = PyPDF2.PdfReader(fh)
            for page in reader.pages:
                text += page.extract_text() or ""

            # If standard text extraction is minimal or failed, attempt OCR
            if not text or len(text.strip()) < 50:
                logger.info(
                    f"Minimal text extracted from {file_id} with PyPDF2, attempting OCR."
                )
                fh.seek(0)  # Reset stream position to read for OCR
                # --- OCR Implementation Needed Here ---
                # This section requires an OCR library (e.g., pytesseract) and potentially a PDF-to-image converter (e.g., pdf2image).
                # Example (conceptual, requires libraries and setup):
                # try:
                #     from PIL import Image
                #     import pytesseract
                #     from pdf2image import convert_from_bytes
                #
                #     images = convert_from_bytes(fh.read())
                #     ocr_text = ""
                #     for i, image in enumerate(images):
                #         ocr_text += pytesseract.image_to_string(image, lang='tha+eng') # Specify languages if needed
                #         ocr_text += "\n" # Add newline between pages
                #     text = ocr_text if ocr_text.strip() else "[OCR could not extract text]"
                #     if text.strip() == "[OCR could not extract text]":
                #          logger.warning(f"OCR failed to extract text for {file_id}.")
                #     else:
                #          logger.info(f"OCR successfully extracted text for {file_id}.")
                #
                # except ImportError:
                #     logger.error("Required libraries for OCR (e.g., pytesseract, pdf2image, PIL) not installed.")
                #     text = "[Error: Required libraries for OCR not installed.]"
                # except Exception as ocr_e:
                #     logger.error(f"Error during OCR processing for {file_id}: {ocr_e}")
                #     text = f"[Error during OCR: {ocr_e}]"

                # Placeholder text if OCR is not implemented or fails
                if (
                    not text
                    or text.strip() == "[OCR could not extract text]"
                    or text.startswith("[Error")
                ):  # Check if OCR attempt was successful or implemented
                    text = "[Content could not be extracted. It might be an image-based PDF without selectable text, and OCR is not fully implemented or failed.]"

        except Exception as e:
            logger.error(f"Error getting PDF content for {file_id} with PyPDF2: {e}")
        return text if text else "[No extractable text in PDF]"

    def _get_excel_content_via_openpyxl(self, file_id: str) -> Dict[str, Any]:
        """
        Gets Excel content by downloading the binary file and reading with openpyxl.

        Args:
            file_id (str): ID of the Excel file.

        Returns:
            Dict[str, Any]: Structured Excel data.
        """
        try:
            # Download the raw Excel file
            request = self.service.files().get_media(fileId=file_id)
            fh = BytesIO()
            downloader = MediaIoBaseDownload(fh, request)

            done = False
            while not done:
                _, done = downloader.next_chunk()

            fh.seek(0)

            # Load the workbook with openpyxl
            # data_only=True retrieves cell values, not formulas
            wb = openpyxl.load_workbook(fh, data_only=True)
            file_metadata = self.service.files().get(fileId=file_id, fields="name").execute()
            file_name = file_metadata.get('name', 'Unknown')

            result = {
                'file_name': file_name,
                'file_id': file_id,
                'sheets': {},
                'summary': f"Excel file '{file_name}' has {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}",
            }

            # Read sample data from each sheet (e.g., first 5 rows, first 5 columns)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_data = []
                # Adjust max_row and max_col as needed to get more sample data
                max_data_row = min(ws.max_row, 5)
                max_data_col = min(ws.max_column, 5)
                
                # Read headers if the first row seems like headers
                headers = []
                if max_data_row >= 1:
                    header_row = [ws.cell(row=1, column=c).value for c in range(1, max_data_col + 1)]
                    # Simple check if first row looks like headers (e.g., not all empty, contains some text)
                    if any(cell is not None and str(cell).strip() for cell in header_row):
                         headers = header_row
                         start_row = 2 # Start data from the second row
                    else:
                         start_row = 1 # Start data from the first row (no header row)

                    # Read sample data rows (excluding header if identified)
                    for r in range(start_row, max_data_row + 1):
                         row_values = [ws.cell(row=r, column=c).value for c in range(1, max_data_col + 1)]
                         rows_data.append(row_values)


                result['sheets'][sheet_name] = {
                    'headers': headers,
                    'sample_data': rows_data,
                    'row_count': ws.max_row,
                    'col_count': ws.max_column
                }

            return result

        except Exception as e:
            logger.error(f"Error reading Excel with openpyxl for {file_id}: {e}")
            return {'error': str(e), 'summary': f'Error reading Excel file: {str(e)}'}

    def _get_other_content(self, file_id: str) -> str:
        """
        Get content from other file types.
        
        Args:
            file_id (str): ID of the file
            
        Returns:
            str: Decoded content from file
        """
        content = self.service.files().get_media(fileId=file_id).execute()
        return content.decode("utf-8", errors="ignore")

    def _get_excel_with_openpyxl(self, file_id: str) -> Dict[str, Any]:
        try:
            # ดาวน์โหลดไฟล์ Excel จาก Google Drive
            request = self.service.files().get_media(fileId=file_id)
            fh = BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            fh.seek(0)

            # โหลดด้วย openpyxl
            wb = openpyxl.load_workbook(fh, data_only=True)
            result = {
                "sheets": {},
                "file_id": file_id,
                "summary": f"Excel file has {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}",
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=2, values_only=True):
                    rows.append(row)
                result["sheets"][sheet_name] = rows

            return result

        except Exception as e:
            logger.error(f"openpyxl error for {file_id}: {e}")
            return {"error": str(e)}

    def view_file(self, file_number: int) -> str:
        """
        View content of a specific file by its number in the list.

        Args:
            file_number (int): Number of the file in the list

        Returns:
            str: File content or error message
        """
        try:
            # Get list of files
            results = (
                self.service.files()
                .list(
                    q=(
                        "name contains 'NDA' or "
                        "name contains 'Non-Disclosure' or "
                        "name contains 'confidential'"
                    ),
                    pageSize=50,
                    fields="files(id,name)",
                )
                .execute()
            )

            files = results.get("files", [])

            if not files:
                return "No NDA documents found in Google Drive."

            if file_number < 1 or file_number > len(files):
                return f"Invalid file number. Please choose a number between 1 and {len(files)}."

            # Get selected file
            selected_file = files[file_number - 1]
            file_id = selected_file["id"]
            file_name = selected_file["name"]

            # Get file content
            content = self.get_file_content(file_id)

            # Check if content is a dictionary (likely from Excel processed by openpyxl)
            if isinstance(content, dict):
                # Format the dictionary content for display
                file_name = content.get("file_name", "Unknown Excel File")
                summary = content.get("summary", "Could not retrieve summary.")
                sheets_data = content.get("sheets", {})

                formatted_content = f"File: {file_name}\n\n{summary}\n\n"

                # Limit the number of sheets shown for brevity
                for sheet_name, sheet_data in list(sheets_data.items())[:3]:
                    formatted_content += f"  --- Sheet: {sheet_name} ---\n"

                    # Check for headers and sample_data keys
                    if 'headers' in sheet_data and sheet_data['headers']:
                        formatted_content += "  Headers: " + ", ".join(map(str, sheet_data['headers'])) + "\n"

                    if 'sample_data' in sheet_data and sheet_data['sample_data']:
                        formatted_content += "  Sample Rows:\n"
                        for i, row_data in enumerate(sheet_data['sample_data']):
                            row_string = ", ".join(map(lambda x: str(x) if x is not None else "", row_data))
                            formatted_content += f"    Row {i + 1}: {row_string}\n"
                    else:
                        formatted_content += "  [Sheet is empty or no sample data available]\n"

                    if 'error' in content and content['error']:
                        formatted_content += f"  Error in data: {content['error']}\n"

                    formatted_content += "\n"  # Add spacing between sheets

                # Return only the formatted content string for Excel, without the outer markdown block
                return formatted_content

            else:
                # Original formatting for string content (for non-Excel files)
                formatted_content = "".join(
                    [line.strip() + "\n" for line in content.splitlines()]
                )
                formatted_content = formatted_content.replace(
                    "\n\n\n", "\n\n"
                )  # Reduce multiple newlines to double newlines

                response = (
                    f"Content of file: {file_name}\n\n" f"```\n{formatted_content}\n```\n"
                )

                return response

        except Exception as e:
            logger.error(f"Error viewing file: {e}")
            return f"Error viewing file: {str(e)}"

    def get_analysis_context(self, nda_content: str) -> str:
        """
        Prepares the context for NDA analysis including the NDA content and loaded guidelines.

        Args:
            nda_content (str): The content of the NDA document to analyze

        Returns:
            str: Formatted context string containing NDA content and guidelines
        """
        context_string = f"NDA Document Content:\n---\n{nda_content}\n---\n\n"

        if self.guidelines:
            context_string += "Loaded Guidelines:\n---\n"
            for name, guideline in self.guidelines.items():
                context_string += (
                    f"Guideline: {guideline.name} (Type: {guideline.type})\n"
                    f"{guideline.content}\n---\n"
                )
            context_string += "\n"
        else:
            context_string += (
                "No specific guidelines loaded. Using general analysis criteria.\n\n"
            )

        # Add static analysis instructions
        context_string += """
ANALYSIS INSTRUCTIONS FOR LLM:
You are an expert in NDA analysis. Analyze the 'NDA Document Content' provided above based on the 'Loaded Guidelines' (if any) and the following general criteria. Prioritize the Loaded Guidelines if they provide more specific instructions than the general criteria.

Follow the ANALYSIS OUTPUT FORMAT strictly.

General Analysis Criteria:
1. Check for essential NDA clauses:
   - Definition of Confidential Information
   - Obligations of Receiving Party
   - Permitted Use
   - Return of Information
   - Term/Duration
   - Remedies
   - Governing Law
   - Signatures

2. Look for potential issues:
   - Overly broad definitions
   - Unreasonable obligations
   - Unclear language
   - Missing essential protections

3. Provide specific recommendations:
   - Clear language suggestions
   - Missing clause templates
   - Compliance improvements

4. When analyzing, provide:
   - A summary of key findings
   - Missing or problematic clauses
   - Specific recommendations
   - Best practices
"""
        return context_string

    def analyze_excel_structure(self, file_id: str) -> Dict[str, Any]:
        """
        Analyze Excel file structure in detail.

        Args:
            file_id (str): ID of the Excel file

        Returns:
            Dict[str, Any]: Detailed analysis of Excel structure
        """
        try:
            excel_data = self._get_excel_content_enhanced(file_id, "")

            if excel_data.get("error"):
                return {"error": excel_data["error"]}

            analysis = {
                "file_info": {
                    "name": excel_data["file_name"],
                    "size": excel_data.get("file_size", 0),
                    "mime_type": excel_data["mime_type"],
                },
                "structure": {},
                "data_types": {},
                "recommendations": [],
            }

            # Analyze CSV data if available
            if excel_data.get("raw_csv"):
                lines = excel_data["raw_csv"].split("\n")
                if len(lines) > 1:
                    headers = [col.strip('"').strip() for col in lines[0].split(",")]

                    # Analyze data types in each column
                    sample_rows = lines[
                        1 : min(6, len(lines))
                    ]  # Sample first 5 data rows

                    for i, header in enumerate(headers):
                        column_samples = []
                        for row in sample_rows:
                            cols = row.split(",")
                            if i < len(cols):
                                column_samples.append(cols[i].strip('"').strip())

                        # Simple data type detection
                        data_type = self._detect_column_type(column_samples)
                        analysis["data_types"][header] = {
                            "detected_type": data_type,
                            "samples": column_samples[:3],  # First 3 samples
                        }

                    analysis["structure"] = {
                        "total_rows": len(lines) - 1,
                        "total_columns": len(headers),
                        "headers": headers,
                        "has_headers": True,
                    }

                    # Generate recommendations
                    if len(headers) > 20:
                        analysis["recommendations"].append(
                            "Large number of columns - consider data normalization"
                        )

                    if len(lines) > 1000:
                        analysis["recommendations"].append(
                            "Large dataset - consider processing in chunks"
                        )

            return analysis

        except Exception as e:
            logger.error(f"Error analyzing Excel structure: {e}")
            return {"error": str(e)}

    def _detect_column_type(self, samples: List[str]) -> str:
        """
        Simple data type detection for Excel columns.

        Args:
            samples (List[str]): Sample values from the column

        Returns:
            str: Detected data type
        """
        if not samples or all(not sample.strip() for sample in samples):
            return "empty"

        # Remove empty samples
        non_empty_samples = [s.strip() for s in samples if s.strip()]

        if not non_empty_samples:
            return "empty"

        # Check for numbers
        numeric_count = 0
        date_count = 0

        for sample in non_empty_samples:
            # Check if numeric
            try:
                float(sample.replace(",", ""))  # Handle comma-separated numbers
                numeric_count += 1
            except ValueError:
                pass

            # Check if date-like
            if any(sep in sample for sep in ["/", "-", "."]):
                if len(sample.split("/")) == 3 or len(sample.split("-")) == 3:
                    date_count += 1

        total_samples = len(non_empty_samples)

        if numeric_count == total_samples:
            return "numeric"
        elif date_count > total_samples * 0.5:
            return "date"
        elif numeric_count > total_samples * 0.5:
            return "mixed_numeric"
        else:
            return "text"


# Create Google Drive Service instance
drive_service = GoogleDriveService()

# Modify functions to use ADK session state
def guidelines_loaded(ctx) -> bool:
    """
    Check if guidelines are loaded using ADK session state.
    Returns True if loaded, False otherwise.
    """
    state = ctx.session().state()
    last_load_time = state.get('guidelines_last_load_time')
    if not last_load_time:
        return False
    
    current_time = datetime.now().timestamp()
    cache_duration = 3600  # 1 hour
    return (current_time - last_load_time) < cache_duration

def ensure_guidelines(tool_context) -> tuple[bool, str]:
    """
    Ensure that guidelines are loaded using ADK session state.
    Returns (True, message) if loaded, (False, error_message) if failed.
    """
    try:
        # Check session state first
        state = tool_context.state
        
        # Load guidelines (pass tool_context)
        status = drive_service.load_guidelines(tool_context)
        
        if "Error" in status:
            raise Exception(status)
        
        # Update session state
        state['guidelines_loaded_status'] = True
        state['guidelines_last_load_time'] = datetime.now().timestamp()
        
        return True, status
    except Exception as e:
        tool_context.state['guidelines_loaded_status'] = False
        tool_context.state['guidelines_load_error'] = str(e)
        return False, str(e)

def handle_nda_analysis(tool_context, nda_content):
    """
    Trigger NDA analysis automatically when NDA content is provided.
    Loads guidelines if needed, prepares context, and delegates to analysis agent.
    Returns the analysis result or error message.
    """
    loaded, msg = ensure_guidelines(tool_context) # Pass tool_context
    if loaded:
        # Prepare context for analysis (must include both NDA content and guidelines)
        analysis_context = drive_service.get_analysis_context(nda_content)
        # Delegate to analysis agent with full context (assuming analysis agent tool also takes context if needed)
        # For now, pass context to the tool call if the tool expects it.
        # NOTE: If get_analysis_context tool itself needs session state, its signature must change.
        # Assuming get_analysis_context doesn't need session state access directly for this call:
        result = nda_analysis_agent.tools[0](analysis_context) # Pass context if tool expects it, otherwise just data
        return result
    else:
        return f"ไม่สามารถโหลด Guidelines ได้: {msg}"

# Example usage in workflow (pseudo-code):
# When user uploads or selects a file, call handle_nda_analysis immediately
# user_uploaded_file = ... (get from UI/event)
# if user_uploaded_file:
#     response = handle_nda_analysis(user_uploaded_file)
#     print(response)  # or send to UI

# --- Define Specialized Sub-Agents ---

# Agent for listing NDA files
nda_listing_agent = LlmAgent(
    name="nda_listing_agent",
    model=MODEL_NAME,
    description="Agent specializing in listing available NDA files from integrated storage like Google Drive.",
    instruction="""You are the NDA Listing Agent. Your SOLE purpose is to list available NDA files from Google Drive when explicitly instructed by the Main Agent or the user. You are a step in the workflow orchestrated by the Main Agent.

Follow these steps:

1.  Receive the delegation or command to list files.
2.  Use the `list_nda_files` tool provided to search for and retrieve the list of NDA documents.
3.  Present the results from the `list_nda_files` tool clearly to the user. Ensure the list is formatted for easy reading, showing file numbers and names.
4.  **Crucially, after presenting the list, your task is complete.** Do not attempt to interpret the user's next action or initiate further steps like viewing files or analysis. The Main Agent will receive the user's next instruction and determine the subsequent action based on their response (e.g., selecting a file number). **Signal completion** after presenting the list.

Available Tools:
- list_nda_files: Use this tool to get a list of NDA documents from Google Drive.

ERROR HANDLING: If the `list_nda_files` tool reports an error, inform the user clearly about the error and state that the listing could not be completed. Suggest they report this issue to the Main Assistant. **Signal completion** after reporting the error, returning control to the Main Agent.
""",
    tools=[drive_service.list_nda_files],
)

# Agent for viewing specific NDA file content
nda_viewing_agent = LlmAgent(
    name="nda_viewing_agent",
    model=MODEL_NAME,
    description="Agent specializing in viewing the content of a specific NDA file by file number from a list provided by the listing agent.",
    instruction="""You are the NDA Viewing Agent. Your SOLE purpose is to retrieve the content of a specific NDA file from Google Drive, identified by a file number provided by the user or the Main Agent. You are a step in the workflow orchestrated by the Main Agent. You will then provide a concise summary of the content instead of the full text, **UNLESS the file is a spreadsheet (like .xls or .xlsx)**.

Follow these steps precisely:

1.  Receive the delegation or command to view a file, along with the specified file number.
2.  Use the `view_file` tool provided with the file number. This tool will return the file content or an error.
3.  **Handle Tool Output**: Carefully examine the result from the `view_file` tool.
    a.  **If the tool reports an Error**: Inform the user clearly about the error message received from the tool and state that the file content could not be retrieved. Suggest they report this issue to the Main Assistant. **Signal completion** after reporting the error, returning control to the Main Agent.
    b.  **If the tool returns Content**: Examine the file name or associated mime type information (if available from the tool output or context, or infer from file extension like .xls or .xlsx). If it indicates a spreadsheet file, follow the instructions in step 4a. Otherwise, follow the instructions in step 4b.

4.  **Process Content Based on File Type**:
    a.  **If it's a Spreadsheet**: Recognize that the agent cannot summarize or analyze spreadsheet content directly using the language model. Output a clear message to the user stating that you have retrieved the content but cannot process it further for summarization or analysis. Present the raw text content received from the `view_file` tool (which should be in a readable format like CSV after recent updates to `view_file`) within a markdown code block (```csv\n...\n```) for readability. **Your task is complete after presenting this message and the raw content.** Do NOT attempt to summarize or delegate for analysis. **Signal completion** after presenting the content, returning control to the Main Agent.

    b.  **If it's NOT a Spreadsheet**: Proceed with summarizing the content. **Do NOT present the full file content directly to the user.** Based on the content received from the `view_file` tool (which should be plain text for non-spreadsheets), generate a concise summary (e.g., mention the type of document, key parties, approximate length, and date if available). Also, confirm that the full content has been successfully retrieved and is ready for analysis by the Analysis Agent.
5.  **Present Summary**: Present ONLY the summary and the confirmation message to the user. **Your task is complete after presenting the summary and confirmation.** Do not attempt to analyze the content or ask further questions about it yourself. The Main Agent or Analysis Agent will handle the next step based on user input. **Signal completion** after presenting the summary, returning control to the Main Agent.

Available Tools:
- view_file: Use this tool with a file number to get the content of a specific NDA document from Google Drive.

ERROR HANDLING: (Covered in step 3a) If the `view_file` tool reports an error (e.g., invalid file number, access issue), inform the user clearly about the error and suggest they report it to the Main Assistant. If the file number is out of the valid range, state the valid range to the user (this check might happen before the tool call or be part of the tool's error reporting - the Agent should handle the tool's error output). **Signal completion** after reporting the error.
""",
    tools=[drive_service.view_file],
)

# Agent for managing guideline documents
nda_guidelines_agent = LlmAgent(
    name="nda_guidelines_agent",
    model=MODEL_NAME,
    description="Agent specializing in loading and displaying guideline documents relevant to NDA analysis.",
    instruction="""You are the NDA Guidelines Agent. Your purpose is to load guideline documents or display their content when explicitly instructed by the Main Agent or the user.

Follow these steps:

1. Receive the command:
   - 'Load guidelines' - Load guidelines (will use cache if available)
   - 'Show guidelines' - Display current guidelines
2. Execute Action:
   - For 'Load guidelines': Use load_guidelines tool
   - For 'Show guidelines': Use get_guideline_content tool
3. Present results to user

Available Tools:
- load_guidelines: Fetch and load guideline documents
- get_guideline_content: Display currently loaded guidelines

Note: Guidelines are automatically cached for 1 hour.
""",
    tools=[drive_service.load_guidelines, drive_service.get_guideline_content],
)

# Agent for comprehensive NDA analysis
nda_analysis_agent = LlmAgent(
    name="nda_analysis_agent",
    model=MODEL_NAME,
    description="Specialized agent for comprehensive analysis and review of Non-Disclosure Agreement documents based on provided content and loaded guidelines, following a specific step-by-step workflow.",
    instruction="""Hello! I am your NDA Analysis Assistant. I perform detailed analysis of NDA documents based on provided content, loaded guidelines, and templates.

My role:
- Analyze NDA document content against guidelines and templates.
- Identify missing/extra sections, violations, and unclear clauses.
- Provide structured analysis results, recommendations, and suggested corrections where applicable.

How I work:
1. I receive the NDA document content and loaded guidelines as input context.
2. I use my knowledge and the provided context to perform a detailed analysis.
3. I structure the analysis results according to the defined output format.

**IMPORTANT: You MUST follow the Output Format below PRECISELY. Use Markdown heavily for clear and beautiful formatting (bolding, lists, code blocks, etc.). Maintain consistent structure and formatting for all sections.**

Output Format:
---
📋 **NDA Analysis Report**

📄 **Document Information**:
* Type: [Document Type]
* Date: YYYY-MM-DD (if found)
* Parties: [Company A] and [Company B] (if found)

📊 **Overall Assessment**:
* **Status**: [Pass/Needs Improvement] (Based on severity and number of issues)
* **Risk Level**: [Low/Medium/High] (Assess overall risk based on critical issues, violations, and missing clauses)
* **Compliance Score**: X/100 (Calculate based on Compliance Score Criteria below. Provide a brief justification for the score.)

**Compliance Score Criteria (Use these weights as a guide):**
- Presence of all Essential Standard Clauses (Definition, Obligations, Term, Governing Law, Signatures): 50 points (Deduct points for each missing clause)
- Absence of Critical Issues/Violations: 35 points (Deduct significant points for each critical issue or violation)
- Clarity and Reasonableness of Clauses: 5 points (Deduct minor points for unclear or potentially unreasonable clauses)
- Absence of Missing Critical Sections (from template comparison): 10 points (Deduct points for each missing critical section)
- Aim for consistent scoring across similar documents.

* **Key Strengths**: (Summarize the well-drafted or compliant aspects)
* **Key Weaknesses**: (Summarize the main areas needing improvement)
* **Overall Conclusion**: (Provide a brief concluding statement summarizing the readiness or required actions for the document)

---

🚨 **Violations Identified**:
(List of violations found, if any. If none, state "None identified." Use bullet points for each violation.)

* **[Violation Type/Summary]**: [Brief description]
  > **Original Text**:
  > ```
  > [Relevant text from the document]
  > ```
  > **Explanation**:
  >[Detailed explanation of the violation]
  >
  > **Legal Basis/Reference**:
  > [Relevant legal principle or guideline reference]
  >
  > **Suggested Rewrite**:
  > ```
  >
  > [Proposed rewritten clause, if applicable, otherwise "N/A"]
  > ```

---

🤔 **Unclear or Unreasonable Clauses**:
(List of unclear/unreasonable clauses found, if any. If none, state "None identified." Use bullet points for each clause.)

* **[Issue Type/Summary]**: [Brief description]
  > **Original Text**:
  > ```
  > [Relevant text from the document]
  > ```
  > **Explanation**:
  > [Detailed explanation of why it is unclear or unreasonable]
  >
  > **Legal Basis/Reference**:
  > [Relevant legal principle or guideline reference]
  >
  > **Suggestion**:
  > [Recommended action or clarification needed]
  >
  > **Suggested Rewrite**:
  > ```
  >
  > [Proposed rewritten clause, if applicable, otherwise "N/A"]
  > ```

---

🔍 **Key Findings**:
(Summary of key points from the analysis, often pulled from Violations, Unclear Clauses, and Missing Sections. If none, state "No significant findings." Use bullet points for findings.)

* [Finding 1]
* [Finding 2]
* [Finding 3]

---

⚠️ **Critical Issues**:
(High-risk issues or critical problems that need addressing. If none, state "None identified." Use bullet points for issues.)

* [Issue 1]
* [Issue 2]
* [Issue 3]

---

💡 **Recommendations**:
(General suggestions for improving the document. Provide at least 2-3 recommendations even if the score is high. Use bullet points for recommendations.)

* [Recommendation 1]
* [Recommendation 2]
* [Recommendation 3]

---

📊 **Completeness Check**:
(Status of essential standard NDA clauses: ✅ Present / ❌ Missing) Use a clear list format.
**IMPORTANT:** Ensure this list is consistent with the 'Missing Sections' below.

* ✅ / ❌ Definition of Confidential Information
* ✅ / ❌ Obligations of Receiving Party
* ✅ / ❌ Permitted Use
* ✅ / ❌ Return of Information
* ✅ / ❌ Term/Duration
* ✅ / ❌ Remedies
* ✅ / ❌ Governing Law
* ✅ / ❌ Signatures
* ✅ / ❌ Authorized Signatories Section

---

🔍 **Template Comparison Results**:

📋 **Missing Sections**:
(Specific critical sections missing based on template comparison. If none, state "None identified." Use bullet points for missing sections.)
**IMPORTANT:** Ensure this list is consistent with the 'Completeness Check' above for missing items.
* [Section 1]
  - Required elements: [List]
  - Impact: [Description]

📋 **Extra Sections**:
(Sections present in the document but not typically found in standard templates. If none, state "None identified." Use bullet points for extra sections.)
* [Section 1]
  - Assessment: [Description]
  - Recommendation: [Keep/Remove/Modify]

---

💡 **Next Steps**:
* To view specific section: "view section [name]"
* To generate report: "generate report"
* To analyze another document: "analyze document"

---

Error Handling:
- If the input context is incomplete, I will report that analysis cannot proceed.
- If I encounter issues during analysis, I will report them in simple language.

Remember to:
- Perform a thorough analysis based on all provided context.
- Use the defined output format strictly.
- Provide clear and actionable recommendations.
- Be specific about missing or extra sections based on template comparison.
- Explain suggested rewrites with clear reasons if applicable.
- Ensure all relevant sections (Violations, Unclear Clauses, Findings, Issues, Recommendations, Completeness Check, Template Comparison, Missing Sections, Extra Sections) are included and populated based on the analysis. Explicitly state "None identified" if a section has no findings.
""",
    tools=[drive_service.get_analysis_context],
)

# Add new agent for Excel operations
excel_operations_agent = LlmAgent(
    name="excel_operations_agent",
    model=MODEL_NAME,
    description="Agent specializing in Excel file operations including detailed analysis and structure inspection.",
    instruction="""You are the Excel Operations Agent. Your purpose is to handle detailed Excel file operations including structure analysis and data inspection.

Available Commands:
1. "analyze excel [file_id]" - Perform detailed structure analysis of an Excel file
2. "inspect excel [file_number]" - Inspect Excel file from the listed files

Follow these steps:

1. When asked to analyze an Excel file:
   - Use the `analyze_excel_structure` tool with the provided file_id
   - Present the analysis results in a clear, structured format
   - Include file info, structure details, data types, and recommendations

2. When asked to inspect an Excel file by number:
   - First get the file list to identify the correct file_id
   - Then use the analysis tools to provide detailed information

Available Tools:
- analyze_excel_structure: Analyze Excel file structure and data types
- list_nda_files: Get list of files to identify file_id from file number

Present results using clear formatting with:
- File information summary
- Data structure overview  
- Column data types
- Sample data preview
- Recommendations for data processing

ERROR HANDLING: If any tool reports an error, inform the user clearly and suggest alternative approaches.
""",
    tools=[drive_service.analyze_excel_structure, drive_service.list_nda_files],
)

# Modify enhanced_nda_assistant instruction to include transfer optimization
enhanced_nda_assistant = LlmAgent(
    name="enhanced_nda_assistant",
    model=MODEL_NAME,
    description="Main Assistant for handling user inquiries about NDAs and managing workflows for tasks like listing, viewing, guidelines, and analysis.",
    instruction="""Introduce yourself to make it easier to use as follows:
Hello! I am your Enhanced NDA Assistant, ready to assist you with Non-Disclosure Agreement documents 🤖📄.

I can help you with Non-Disclosure Agreements. Here is what I can do:

1.  List available NDA files from Google Drive.
2.  View the content of a specific NDA file by its number.
3.  Analyze an NDA document and provide a detailed report based on guidelines.
4.  Show the loaded guidelines used for analysis.

Just let me know what you'd like to do! 👍

WORKFLOW FOR DOCUMENT ANALYSIS:
- When a user asks to analyze an NDA document (by providing content or referring to previously viewed content), I will orchestrate the analysis process through these steps:
  1. **Set Legal System:** I will automatically set the legal system to Thai Rules in the session state.
     - Set `legal_system` key in session state to "Thai Rules"
     - No need to ask the user about legal system choice
  2. **Load Guidelines:** I will delegate the task of loading guidelines to the `nda_guidelines_agent` by instructing it to "Load guidelines". I will wait for the result.
  3. **Check Guideline Status:** I will check the result from the `nda_guidelines_agent`. If guideline loading failed (indicated by an error message), I will report the error to the user and inform them that analysis cannot proceed.
  4. **Prepare Analysis Context:** If guidelines loaded successfully, I will use the `drive_service.get_analysis_context` tool, passing the NDA document content to prepare the full context for the analysis agent (this context will include both the NDA and the loaded guidelines).
  5. **Perform Analysis:** I will delegate the actual analysis task to the `nda_analysis_agent`. I will pass the analysis context prepared in the previous step as the input to the `nda_analysis_agent`. I will wait for the analysis report from the `nda_analysis_agent`.
  6. **Present Results:** I will present the final analysis report received from the `nda_analysis_agent` to the user using clear formatting.

USER EXPERIENCE:
- You do NOT need to ask the user about guideline loading before initiating the analysis; I will handle that step within the workflow.
- Always provide clear, concise feedback and next steps after each action.
- If an error occurs in any delegated task or tool call, I will acknowledge it, inform the user, and guide them on potential next steps.

COMMANDS I UNDERSTAND:
- "Analyze this document" (Initiates the analysis workflow)
- "List NDA files" (Delegates to nda_listing_agent)
- "View file [number]" (Context-dependent, delegates to nda_viewing_agent)
- "Show guidelines" (Delegates to nda_guidelines_agent)
- "Load guidelines" (Delegates to nda_guidelines_agent for manual loading)
- "Analyze excel [file_id]" or "Inspect excel [file_number]" (Delegates to excel_operations_agent)
- General NDA questions
- "Exit"

ERROR HANDLING:
- If a delegated agent or a tool call within a delegated agent reports an error, I will acknowledge it, inform the user, and guide them on potential next steps.

AGENT TRANSFER INSTRUCTIONS:
- Use the built-in transfer_to_agent function to delegate tasks to specialized agents
- For listing files: transfer_to_agent(agent_name='nda_listing_agent')
- For viewing files: transfer_to_agent(agent_name='nda_viewing_agent')
- For analysis: transfer_to_agent(agent_name='nda_analysis_agent')
- For guidelines: transfer_to_agent(agent_name='nda_guidelines_agent')
- For Excel operations: transfer_to_agent(agent_name='excel_operations_agent')

- Batch related operations when possible to minimize transfers
- Maintain context between transfers using session state
- Return control to main agent after task completion
""",
    tools=[drive_service.get_analysis_context],
    sub_agents=[
        nda_listing_agent,
        nda_viewing_agent,
        nda_analysis_agent,
        excel_operations_agent,
        nda_guidelines_agent,
    ],
)

# Set as root agent
root_agent = enhanced_nda_assistant